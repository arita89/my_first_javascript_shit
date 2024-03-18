# services.py

import csv
import yaml
import pandas as pd
from tempfile import NamedTemporaryFile
import os

# to create and format excel file
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

def apply_sheet_colors(writer, sheet_name, color_code):
    """
    Applies color to sheet tabs using openpyxl.

    Parameters:
    - writer: The pd.ExcelWriter instance
    - sheet_name: The name of the sheet to color
    - color_code: The hex color code (expected without '#')
    """
    workbook = writer.book
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet.sheet_properties.tabColor = color_code

def apply_conditional_formatting(worksheet, column_letter, validation_type, validation_details):
    """
    Applies conditional formatting: yellow for empty cells, and red if cells do not meet the validation criteria.

    Parameters:
    - worksheet: The worksheet object to apply conditional formatting to.
    - column_letter: The letter of the column to apply formatting.
    - validation_type: The type of validation ('range', 'list', etc.).
    - validation_details: The details for the validation rule.
    """
    # Define fill colors
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Apply yellow fill for empty cells
    worksheet.conditional_formatting.add(f'{column_letter}2:{column_letter}1048576',
                                         CellIsRule(operator='equal', formula=['""'], fill=yellow_fill))

    # Apply red fill based on validation criteria
    if validation_type == 'range':
        # For a range validation, e.g., '>0', we construct a formula to check this
        # Note: This is a simplistic approach and may need adjustment for complex validations
        start, end = validation_details.strip('[]').split('-')
        formula = f'OR({column_letter}2<{start}, {column_letter}2>{end})'
        worksheet.conditional_formatting.add(f'{column_letter}2:{column_letter}1048576',
                                             FormulaRule(formula=[formula], fill=red_fill, stopIfTrue=True))

    elif validation_type == 'min':
        # Extract the number after '>'
        minimum = validation_details.split('>')[1]
        formula = f"{column_letter}2<={minimum}"
        worksheet.conditional_formatting.add(f'{column_letter}2:{column_letter}1048576',
                                             FormulaRule(formula=[formula], fill=red_fill, stopIfTrue=True))
    elif validation_type == 'max':
        # Extract the number after '<'
        maximum = validation_details.split('<')[1]
        formula = f"{column_letter}2>={maximum}"
        worksheet.conditional_formatting.add(f'{column_letter}2:{column_letter}1048576',
                                             FormulaRule(formula=[formula], fill=red_fill, stopIfTrue=True))
        
    elif validation_type == 'list':
        # For list validations, checking if the cell value is not in the list
        # This requires a bit of creativity as direct list checking isn't straightforward in conditional formatting
        values = validation_details.split(',')
        formula = 'AND(' + ', '.join([f'{column_letter}2<>"{value}"' for value in values]) + ')'
        worksheet.conditional_formatting.add(f'{column_letter}2:{column_letter}1048576',
                                             FormulaRule(formula=[formula], fill=red_fill, stopIfTrue=True))
        
def apply_validations(worksheet, column_letter, validation_type, validation_details):
    """
    Applies data validation to a column based on the specified validation type and details.

    Parameters:
    - worksheet: The openpyxl worksheet object to which the validation will be applied.
    - column_letter: The letter of the column to apply the validation to.
    - validation_type: The type of validation to apply ('range', 'list', etc.).
    - validation_details: The specific details for the validation (e.g., the range or list of values).
    """
    if validation_type and validation_type not in ['na', ' ', None]:
        if validation_type == 'range':
            start, end = validation_details.split('-')
            dv = DataValidation(type="whole", operator="between", formula1=start, formula2=end, showInputMessage=True)
        elif validation_type == 'list':
            dv = DataValidation(type="list", formula1=f'"{validation_details}"', showDropDown=True, showInputMessage=True)
        else:
            return  # Placeholder for other validation types
        
        dv.add(f'{column_letter}2:{column_letter}1048576')
        worksheet.add_data_validation(dv)   

def csv_to_yaml(csv_content, delimiter=';'):
    """
    Converts CSV content to a YAML string. The CSV is expected to define sheets, columns, and data validations.

    Parameters:
    - csv_content: The content of the CSV file as a string.
    - delimiter: The delimiter used in the CSV file (default is ';').

    Returns:
    - A string containing the YAML representation of the CSV content.
    """
    csv_reader = csv.DictReader(csv_content.splitlines(), delimiter=delimiter)
    yaml_data = {'sheets': []}
    current_sheet_name = ""
    sheet_structure = {}

    for row in csv_reader:
        if not row['sheet'].strip():  # Skip rows with blank sheet names, in case of trailing white spaces on actually empty rows
            continue
        if current_sheet_name != row['sheet'].strip().lower():  # Normalize sheet names
            if current_sheet_name:
                yaml_data['sheets'].append(sheet_structure)
            current_sheet_name = row['sheet'].strip().lower()  # Normalize sheet names
            sheet_structure = {
                'name': current_sheet_name,
                'color': row.get('sheet_color', 'FFFFFF'),  # Assume white as default color if not specified
                'is_optional': row['sheet_is_optional'].lower() == 'true',
                'responsibility': row['responsibility'],
                'columns': []
            }

        # Prepare column details, handling "na" or missing validation info
        column_details = {key.replace("column_", "", 1): (value if value.lower() != 'na' else None) for key, value in row.items() if key.startswith('column_')}
        sheet_structure['columns'].append(column_details)

    yaml_data['sheets'].append(sheet_structure)
    return yaml.dump(yaml_data, sort_keys=False, default_flow_style=False, allow_unicode=True)


def yaml_to_excel(yaml_str):
    """
    Converts YAML string data into an Excel file with sheets, columns, and validations as defined in the YAML.

    Parameters:
    - yaml_str: The YAML string containing the sheet definitions and validations.

    Returns:
    - The path to the generated Excel file.
    """
    data = yaml.safe_load(yaml_str)
    temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_file_path = temp_file.name
    temp_file.close()
    
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        for sheet in data['sheets']:
            sheet_name = sheet['name'][:31].strip()  # Ensure valid Excel sheet name
            df = pd.DataFrame(columns=[col['name_internal'] for col in sheet['columns']])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            
            # Apply sheet tab color
            if 'color' in sheet:
                color_code = sheet['color'].lstrip('#')
                apply_sheet_colors(writer, sheet_name, color_code)
                
            # Apply validations
            for idx, col in enumerate(sheet['columns'], start=1):
                column_letter = get_column_letter(idx)
                if 'data_validation_type' in col and col['data_validation_type'] not in ['na', ' ', None] and 'data_validation_details' in col:
                    apply_validations(writer.sheets[sheet_name], column_letter, col['data_validation_type'], col['data_validation_details'])
                
                # Apply conditional formatting based on your criteria
                # Adjust this line to include the validation type and details
                apply_conditional_formatting(worksheet, column_letter, 
                                             col.get('data_validation_type'), 
                                             col.get('data_validation_details'))

    return excel_file_path

def append_to_excel_as_hidden_locked(source_path, target_path, sheet_name, password):
    # Determine if the source is CSV or Excel and load it into a DataFrame
    if source_path.endswith('.csv'):
        df = pd.read_csv(source_path)
    else:  # For Excel files
        df = pd.read_excel(source_path)
    
    #print (df) 

    # Load the target workbook
    workbook = load_workbook(target_path)

    # Check if the sheet name already exists
    if sheet_name in workbook.sheetnames:
        msg = f"Sheet name {sheet_name} already exists in target workbook."
        print (msg)
        return msg

    # Create a new sheet in the workbook
    sheet = workbook.create_sheet(title=sheet_name)

    # Fill the new sheet with data from the DataFrame
    for row in dataframe_to_rows(df, index=False, header=True):
        print (row)
        sheet.append(row)

    # Hide and Protect the sheet
    sheet.sheet_state = 'hidden'  # Hide the sheet
    sheet.protection.set_password(password)

    # Save the workbook
    workbook.save(target_path)
    msg = "Success! Source appended to target workbook as hidden, locked sheet."
    return msg


def unlock_hidden_sheets(excel_file_path, password):
    workbook = load_workbook(excel_file_path)
    unlocked_sheets = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet.sheet_state == 'hidden':
            try:
                sheet.protection.disable()
                sheet.sheet_state = 'visible'
                unlocked_sheets.append(sheet_name)
            except Exception as e:
                # Log or handle the fact that some sheets might not be unlocked due to incorrect password or other issues
                return e
    workbook.save(excel_file_path)
    return unlocked_sheets


# Function to clean up the generated file
def cleanup_file(file_path):
    os.remove(file_path)


