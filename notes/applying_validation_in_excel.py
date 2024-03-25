import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Load CSV into DataFrame
df = pd.read_csv("validation_rules.csv")

# Load Excel workbook
wb = load_workbook("target_workbook.xlsx")

for index, row in df.iterrows():
    sheet_name = row["sheet"]
    if sheet_name not in wb.sheetnames and row["sheet_is_optional"].lower() == "yes":
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    column = row["column_name"]
    val_type = row["val_type"]
    formula1 = row["val_formula1"]
    formula2 = row["val_formula2"] if pd.notnull(row["val_formula2"]) else None

    # Adjust list items format if val_type is 'list'
    if val_type == "list" and pd.notnull(formula1):
        formula1 = (
            '"'
            + ",".join(
                [f'"{item.strip()}"' for item in formula1.strip("[]").split(",")]
            )
            + '"'
        )

    # Create DataValidation object
    dv = DataValidation(
        type=val_type,
        operator=row["val_operator"] if pd.notnull(row["val_operator"]) else None,
        formula1=formula1,
        formula2=formula2,
        showDropDown=(val_type == "list"),
    )

    # Apply to specified range or entire column if no range is specified
    apply_range = (
        row["val_range"]
        if pd.notnull(row["val_range"])
        else f"{column}2:{column}1048576"
    )
    dv.add(apply_range)
    ws.add_data_validation(dv)

wb.save("updated_target_workbook.xlsx")


from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active

# List validation
dv_list = DataValidation(type="list", formula1='"Dog,Cat,Bird"', showDropDown=True)
ws.add_data_validation(dv_list)
dv_list.add(ws["A1"])  # Apply to cell A1

# Whole number validation
dv_whole = DataValidation(
    type="whole", operator="between", formula1="1", formula2="100"
)
ws.add_data_validation(dv_whole)
dv_whole.add(ws["B1:B5"])  # Apply to range B1:B5

# Decimal validation
dv_decimal = DataValidation(
    type="decimal", operator="between", formula1="0.1", formula2="9.9"
)
ws.add_data_validation(dv_decimal)
dv_decimal.add("C1:C5")  # Also supports range strings

# Date validation
dv_date = DataValidation(
    type="date", operator="between", formula1="2020-01-01", formula2="2024-12-31"
)
ws.add_data_validation(dv_date)
dv_date.add("D1:D5")

# Time validation
dv_time = DataValidation(
    type="time", operator="between", formula1="06:00", formula2="18:00"
)
ws.add_data_validation(dv_time)
dv_time.add("E1:E5")

# Text length validation
dv_text_length = DataValidation(type="textLength", operator="lessThan", formula1="10")
ws.add_data_validation(dv_text_length)
dv_text_length.add("F1:F5")

# Custom formula validation
dv_custom = DataValidation(type="custom", formula1="=LEN(A1)=5")
ws.add_data_validation(dv_custom)
dv_custom.add("G1:G5")

wb.save("data_validation_demo.xlsx")
